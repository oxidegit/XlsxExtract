from openpyxl import load_workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
wb = load_workbook('统计结果.xlsx')
for sheet in wb:
    if (sheet.sheet_state != 'hidden'):
        for row in sheet.iter_rows(2, sheet.max_row, 6, 17):
            for cell in row:
                if (cell.value is None):
                    cell.value=0

wb.save('统计结果.xlsx')