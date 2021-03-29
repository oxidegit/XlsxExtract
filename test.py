####################合并单元格&& 日期格式设置
from openpyxl import load_workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment
import re
def testTime():
    wb = load_workbook('2015-2019电影数据（出品方、观影场次人次、月票房、导演、编剧、主演、发行公司）.xlsx')
    wbTest = load_workbook('text.xlsx')
    sheet = wbTest['test']
    a = sheet['D']
    sheet['B1'].fill = PatternFill("solid", fgColor="66CC33")
    wbTest['test']['A1'] = sheet['C41'].value

    #wbTest['test']['A1'].number_format = 'm月d'
    #wbTest['test']['A2'] = None
    wbTest.save('text.xlsx')
def testDao():#提取导演数据
    wb = load_workbook('2015-2019电影数据（出品方、观影场次人次、月票房、导演、编剧、主演、发行公司）.xlsx')
    sheet = wb['2019-1']
    for cell in sheet[2]:
        if (cell.value == '导演'):
            print(cell.column_letter)
class Data:
    def __init__(self):
        self.hello = 'zouc'
    def get(self):
        print(self.hello)
class FilmAward:
    def __init__(self, fileName):
        self.awardXlsx = load_workbook(fileName)
        ws = self.awardXlsx.active
        self.awardData = {}
        for row in ws.iter_rows(3, ws.max_row, 1, ws.max_column):
            p = re.compile(r'[《](.*?)[》]', re.S)  # 最小匹配
            year = int(row[0].value)
            self.awardData[year] = {}
            self.awardData[year]['foreign'] = []
            self.awardData[year]['china'] = []
            for cell in row:
                if ((cell.column>=2 and cell.column<=6) and cell.value):
                    self.awardData[year]['foreign'] = self.awardData[year]['foreign']+re.findall(p, cell.value)
                if(cell.column>6 and cell.value):
                    self.awardData[year]['china'] = self.awardData[year]['china']+re.findall(p, cell.value)
if __name__ == '__main__':
    # string = '《zou》, 《邹成》'
    # p1 = re.compile(r'[《](.*?)[》]', re.S)  # 最小匹配
    # print(re.findall(p1, string))
    # film = FilmAward('电影节奖项及获奖影片整理.xlsx')
    #testDao()
    testTime()
