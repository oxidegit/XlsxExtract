#此类用于提取数据
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment
from openpyxl import Workbook
import re
import os

class FilmAward:
    def __init__(self, fileName):
        self.awardXlsx = load_workbook(fileName)
        ws = self.awardXlsx.active
        self.awardData = {}
        for row in ws.iter_rows(3, ws.max_row, 1, ws.max_column):
            p = re.compile(r'[《](.*?)[》]', re.S)  # 最小匹配
            year = str(row[0].value)
            self.awardData[year] = {}
            self.awardData[year]['foreign'] = []
            self.awardData[year]['china'] = []
            for cell in row:
                if ((cell.column>=2 and cell.column<=6) and cell.value):
                    self.awardData[year]['foreign'] = self.awardData[year]['foreign']+re.findall(p, cell.value)
                if(cell.column>6 and cell.value):
                    self.awardData[year]['china'] = self.awardData[year]['china']+re.findall(p, cell.value)
#电影实体类
class Film:
    def __init__(self, filmInfo, month):
        self.date = filmInfo[0]
        self.name = filmInfo[1]
        self.country = filmInfo[2]
        self.company = filmInfo[3]#出品方
        self.money = [0 for i in range(12)]
        self.count = filmInfo[4] #场次
        self.peopleNum = filmInfo[5]
        self.money[month] = filmInfo[6]
        self.lastMoney = filmInfo[7]
        self.director = filmInfo[8]
        self.writer = filmInfo[9]
        self.actor = filmInfo[10]
        self.province = None
        self.city = None
        self.totalMoney = filmInfo[6]
        self.foreignAward = '否'
        self.chinaAward = '否'
    def setDate(self, date):
        self.date = date
    def setCompany(self, company):
        self.company = company
    def setForeign(self, state):
        self.foreignAward = state
    def setChina(self, state):
        self.chinaAward = state
    def output(self):
        print(self.date, self.name, self.country)

class XlsxResult:
    def __init__(self, resultName):
        self.name = resultName
        if (os.path.exists(resultName)):
            os.remove(resultName)

        self.result = Workbook()
        self.result.save(resultName)
    def modifyStyle(self, ws):
        for i in range(27):
            ws.column_dimensions[get_column_letter(i+1)].width = 15
        ws.column_dimensions['A'].width = 31
    def createSheet(self, sheetName):
        if (sheetName in self.result.sheetnames):
            self.ws = self.result[sheetName]
        else:
            self.ws = self.result.create_sheet(sheetName)
            self.ws.append(
                ['影片名称', '国别', '省', '市', '票房收入', '一月票房收入', '二月票房收入', '三月票房收入', '四月票房收入', '五月票房收入', '六月票房收入',
                 '七月票房收入','八月票房收入', '九月票房收入', '十月票房收入', '十一月票房收入', '十二月票房收入',
                 '上映时间', '票房截止时间（下架）', '出品方', '场次', '人次', '导演', '编剧', '主演', '是否获国际奖',
                 '是否获国内奖'])
            for cell in self.ws[1]:
                cell.fill = PatternFill("solid", fgColor="66CC33")
            self.modifyStyle(self.ws)
        return self.ws
    def save(self):
        self.result.save(self.name)
    def write(self, film):
        maxRow = self.ws.max_row+1
        self.ws.append([film.date])
        self.ws.cell(row=maxRow, column=1, value=film.name)
        self.ws.cell(row=maxRow, column=2, value=film.country)
        self.ws.cell(row=maxRow, column=3, value=film.province)
        self.ws.cell(row=maxRow, column=4, value=film.city)
        self.ws.cell(row=maxRow, column=5, value=film.totalMoney)
        for i in range(12):
            self.ws.cell(row=maxRow, column=6+i, value=film.money[i])
        self.ws.cell(row=maxRow, column=18, value=film.date)
        self.ws.cell(row=maxRow, column=18).number_format = 'm月d'
        self.ws.cell(row=maxRow, column=19, value=film.lastMoney)
        self.ws.cell(row=maxRow, column=20, value=film.company)
        self.ws.cell(row=maxRow, column=21, value=film.count)
        self.ws.cell(row=maxRow, column=22, value=film.peopleNum)
        self.ws.cell(row=maxRow, column=23, value=film.director)
        self.ws.cell(row=maxRow, column=24, value=film.writer)
        self.ws.cell(row=maxRow, column=25, value=film.actor)
        self.ws.cell(row=maxRow, column=26, value=film.foreignAward)
        self.ws.cell(row=maxRow, column=27, value=film.foreignAward)
        #self.result.save(self.name)
class XlsxData:
    def __init__(self, xlsxName):
        self.data = load_workbook(xlsxName)
    def findIndex(self, sheet):#找影片所在的列
        for cell in sheet[1]:
            if (cell.value == '上映日期'):
                return cell.column
    def extractData(self, XlsxResult, award):
        for sheet in self.data:
            if (sheet.sheet_state != 'hidden' and sheet.title != 'top150榜单'):
                year = sheet.title.split('-', 1)[0]
                print(year)
                month = sheet.title.split('-', 1)[1]
                ws = XlsxResult.createSheet(year + '年')
                XlsxResult.result['Sheet'].sheet_state = 'hidden'
                column = self.findIndex(sheet)
                currentDate = None

                begin_row = 3
                end_row = sheet.max_row
                current_row = begin_row
                while (current_row <= end_row and sheet.cell(row=current_row, column=column+1).value):
                    filmInfo = []
                    #判断当前电影所占单元格数
                    if (sheet.cell(row=current_row, column=column).value):
                        pre_date = sheet.cell(row=current_row, column=column).value #上一个有效日期
                    len = 1
                    if (sheet.cell(row=current_row, column=column+3).value):
                        company = sheet.cell(row=current_row, column=column+3).value
                    else:
                        company=''
                    while ((current_row+len<=end_row) and (not sheet.cell(row=current_row+len, column=column-1).value)):
                        len = len + 1
                        if ((sheet.cell(row=current_row+len, column=column+3).value) and sheet.cell(row=current_row+len, column=column+3).value):
                            company = company + ',' + sheet.cell(row=current_row+len, column=column+3).value
                    #读取current行数据
                    for row in sheet.iter_rows(current_row, current_row, column, column+11):
                        for cell in row:
                            filmInfo.append(cell.value)
                    film = Film(filmInfo, int(month) - 1)
                    film.setCompany(company)
                    film.setDate(pre_date)
                    if (sheet.cell(row=current_row, column=column+1).value in award[year]['foreign']):
                        film.setForeign('是')
                    if (sheet.cell(row=current_row, column=column+1).value in award[year]['china']):
                        film.setForeign('是')
                    current_row = current_row + len
                    XlsxResult.write(film)

if __name__ == '__main__':
    award = FilmAward('电影节奖项及获奖影片整理.xlsx')
    res = XlsxResult('统计结果.xlsx')
    data = XlsxData('2015-2019电影数据（出品方、观影场次人次、月票房、导演、编剧、主演、发行公司）.xlsx')
    data.extractData(res, award.awardData)
    res.save()