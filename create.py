from openpyxl import Workbook
from openpyxl import load_workbook
wb = Workbook()
ws = wb.create_sheet("test")
ws.append(['影片名称', '国别', '省', '市', '票房收入累计', '一月票房收入', '二月票房收入', '三月票房收入', '四月票房收入', '五月票房收入', '六月票房收入', '七月票房收入', '八月票房收入', '九月票房收入', '十月票房收入', '十一月票房收入', '十二月票房收入'])
ws.column_dimensions['A'].width = 31
ws.column_dimensions['B'].width = 6
for i in range(0, 15):
    ws.column_dimensions[chr(67+i)].width = 15
row = ws.max_row + 1
ws.cell(row=row, column=1, value='name.value')
ws.cell(row=row, column=2, value='country.value')
ws.cell(row=row, column=5, value='money.value')
wb.save('text.xlsx')
