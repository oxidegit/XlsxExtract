from openpyxl import load_workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def getFilmIndex(fileName, films):
    index = 1;
    for film in films:
        if (film.value == fileName):
            return index
        index = index+1
    return -1

def findIndex(row):
    index = 1;
    for cell in row:
        if (cell.value == '影片'):
            name = index
        if (cell.value == '国别/地区'):
            country = index
        if (cell.value == '票房(万元)'):
            money = index
        index = index + 1
    return name, country, money

def create(name):
    workbook = load_workbook('text.xlsx')
    ws1 = workbook.create_sheet(name)
    ws1.append(['影片名称', '国别', '省', '市', '票房收入累计', '一月票房收入', '二月票房收入', '三月票房收入', '四月票房收入', '五月票房收入', '六月票房收入', '七月票房收入',
                '八月票房收入', '九月票房收入', '十月票房收入', '十一月票房收入', '十二月票房收入'])
    ws1.column_dimensions['A'].width = 31
    ws1.column_dimensions['B'].width = 6
    for i in range(0, 15):
        ws1.column_dimensions[chr(67 + i)].width = 15
    workbook.save('text.xlsx')
    return ws1

if __name__ == '__main__':
    workbook = load_workbook('text.xlsx')
    wb = load_workbook('2015-2019电影数据（出品方、观影场次人次、月票房、导演、编剧、主演、发行公司）.xlsx')
    for sheet in wb:
        if (sheet.sheet_state != 'hidden' and sheet.title!='top150榜单'):
            year = sheet.title.split('-', 1)[0]
            print(year)
            month = sheet.title.split('-', 1)[1]
            #ws = create(year+'年')
            if (year+'年' in workbook.sheetnames):
                ws = workbook[year + '年']
            else:
                ws = workbook.create_sheet(year + '年')
                ws.append(
                    ['影片名称', '国别', '省', '市', '票房收入累计', '一月票房收入', '二月票房收入', '三月票房收入', '四月票房收入', '五月票房收入', '六月票房收入',
                     '七月票房收入',
                     '八月票房收入', '九月票房收入', '十月票房收入', '十一月票房收入', '十二月票房收入'])
                ws.column_dimensions['A'].width = 31
                ws.column_dimensions['B'].width = 6
                for i in range(0, 15):
                    ws.column_dimensions[chr(67 + i)].width = 15

            index = findIndex(sheet[1])
            filmName = sheet[get_column_letter(index[0])][1:]
            filmCountry = sheet[get_column_letter(index[1])][1:]
            filmMoney = sheet[get_column_letter(index[2])][1:]

            for name, country, money in zip(filmName, filmCountry, filmMoney):
                if name.value:

                    rowLocation = getFilmIndex(name.value, ws[1])
                    if (rowLocation == -1):
                        maxRow = ws.max_row+1
                        ws.cell(row=maxRow, column=1, value=name.value)
                        ws.cell(row=maxRow, column=2, value=country.value)
                        ws.cell(row=maxRow, column=5+int(month), value=money.value)
                        total = ws.cell(row=maxRow, column=5)
                        if (total.value is None):
                            total.value = money.value
                        else:
                            total.value = int(total.value)+int(money.value)
                        workbook.save('text.xlsx')
                    else:
                        ws.cell(row=rowLocation, column=1, value=name.value)
                        ws.cell(row=rowLocation, column=2, value=country.value)
                        ws.cell(row=rowLocation, column=5 + int(month), value=money.value)
                        total = ws.cell(row=rowLocation, column=5)
                        if (total.value is None):
                            total.value = money.value
                        else:
                            total.value = int(total.value) + int(money.value)
                        workbook.save('text.xlsx')
    workbook.save('text.xlsx')







