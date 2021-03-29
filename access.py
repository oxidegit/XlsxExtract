from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def findIndex(row):
    index = 1;
    for cell in row:
        if (cell.value == '影片'):
            name = index
        if (cell.value == '国别/地区'):
            country = index
        if (cell.value == '票房(万元)'):
            money = index
        index = index + 1
    return name, country, money
def getFilmIndex(fileName, films):
    index = 1;
    for film in films:
        if (film.value == fileName):
            return index
        index = index+1
    return -1

wb = load_workbook('text.xlsx')
ws = wb['2019年']


